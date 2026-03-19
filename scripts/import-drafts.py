"""
Import draft history from Excel files into Supabase.

Usage:
    python3 scripts/import-drafts.py

Reads all draft Excel files from the project root and imports picks into
the draft_picks and players tables.

File formats vary by year:
  2011-2019: "Master Draft List" tab — pick#, owner, player, position, ...
  2020: "Draft by Round" tab — pick, team_name, player, team, position, keeper
  2021: "Draft by Round" in 2021-2022 file — overall, round/pick, owner, player, team, position, keeper
  2022: "2022 Draft by Round wKeepers" — pick, round/pick, team, owner, player, team, position, keeper
"""

import os
import re
import json
from supabase import create_client
import openpyxl

# ============================================================
# CONFIG
# ============================================================

SUPABASE_URL = "https://untdbanhuzzibjelwpxp.supabase.co"
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InVudGRiYW5odXp6aWJqZWx3cHhwIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3Mzg0ODc2OSwiZXhwIjoyMDg5NDI0NzY5fQ.X5szKSmzmWvTKEXsBHlwNrOGh7PgMCAistQIzWGJ4Zk")
BASE_DIR = "/Users/alexaltman/fantasy-football"

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

# ============================================================
# OWNER NAME NORMALIZATION
# ============================================================

# Maps first names / short names used in draft sheets → full owner names in DB
NAME_MAP = {
    "alex": "Alex Altman",
    "joel": "Joel Oubre",
    "kevin": "Kevin Whitlock",
    "bill": "Bill Kling",
    "ryan": "Ryan Parrilla",
    "kelly": "Kelly Mann",
    "justin": "Justin Choy",
    "ed": "Ed Lang",
    "sal": "Sal Singh",
    "navi": "Navi Singh",
    "aaron": "Aaron Schwartz",
    "marcus": "Marcus Moore",
    "jason": "Jason McCartney",
    "lance": "Lance Michihira",
    "matt": "Matt B",
}

# Team name → owner name (for FanDraft format)
TEAM_NAME_MAP = {
    "return of the goat": "Alex Altman",
    "three peat": "Alex Altman",
    "joe burrow your picks next year": "Joel Oubre",
    "veronica v. vigilante's": "Joel Oubre",
    "veronica v vigilante's": "Joel Oubre",
    "lela's all-stars": "Kevin Whitlock",
    "lela's all-stars": "Kevin Whitlock",
    "katie 'more' organ": "Bill Kling",
    "titties titties": "Bill Kling",
    "lana rhoades to victory": "Ryan Parrilla",
    "hub city rage": "Kelly Mann",
    "hub city black rage": "Kelly Mann",
    "throw mia some damn balls!!!": "Kelly Mann",
    "big angry tuna": "Kelly Mann",
    "gabriella's a fox": "Justin Choy",
    "gabriella's a fox": "Justin Choy",
    "two mannings one cup": "Ed Lang",
    "new world order": "Ed Lang",
    "alexis mi_amore": "Sal Singh",
    "london keyes to the game": "Navi Singh",
    "kya tropic": "Aaron Schwartz",
    "skull phucker": "Marcus Moore",
    "blood wings": "Marcus Moore",
    "claw dynasty": "Marcus Moore",
    "claw  dynasty": "Marcus Moore",
    "smokedcheddathaassgettah": "Marcus Moore",
    "saquon torches bush burrow": "Joel Oubre",
    "trey burrows in mia khalifa": "Joel Oubre",
    "diggler's thunderstick": "Jason McCartney",
    "diggler's  thunderstick": "Jason McCartney",
    "dame mas gas-alina": "Lance Michihira",
    "dame mas gas alina": "Lance Michihira",
    # FanDraft CSV variations (2022-2025)
    "alexis mi amore": "Sal Singh",
    "alexis mi_amore": "Sal Singh",
    "gabriellas a fox": "Justin Choy",
    "gabriella's a fox": "Justin Choy",
    "smoked chedda the ass getta": "Marcus Moore",
    "lana roades to victory": "Ryan Parrilla",
    "lana rhoades to victory": "Ryan Parrilla",
    "diggler's thunderstick": "Jason McCartney",
    "digglers thunderstick": "Jason McCartney",
    "joe burrow your picks next year": "Joel Oubre",
    "joe burrow your pick's next year": "Joel Oubre",
    "saquon torches bush burrow": "Joel Oubre",
    "return of goat": "Alex Altman",
    "return of the goat": "Alex Altman",
    "lela's all stars": "Kevin Whitlock",
    "lela's all-stars": "Kevin Whitlock",
}


def normalize_owner(raw: str, year: int) -> str:
    """Convert a raw owner string from a draft sheet to a canonical owner name."""
    if not raw:
        return "Unknown"

    clean = raw.strip()

    # Remove trade annotations like "(from Kelly)" or "(via Joel)" or trailing *
    clean = re.sub(r'\s*\(.*?\)', '', clean)
    clean = re.sub(r'\*$', '', clean).strip()

    # Try first-name lookup
    key = clean.lower().strip()
    if key in NAME_MAP:
        return NAME_MAP[key]

    # Try team name lookup
    if key in TEAM_NAME_MAP:
        return TEAM_NAME_MAP[key]

    # Partial match on team names
    for team_key, owner in TEAM_NAME_MAP.items():
        if team_key in key or key in team_key:
            return owner

    print(f"  ⚠️  Unknown owner: '{raw}' (year {year})")
    return raw


def parse_traded_from(raw: str) -> str:
    """Extract the original owner from trade annotations like 'Alex (from Kelly)'."""
    match = re.search(r'\((?:from|via)\s+(\w+)\)', raw, re.IGNORECASE)
    if match:
        return match.group(1).strip()
    return None


def parse_keeper(raw) -> int:
    """Parse keeper status. Returns keeper year (1-4) or None."""
    if not raw:
        return None
    s = str(raw).strip().upper()
    match = re.match(r'K(\d)', s)
    if match:
        return int(match.group(1))
    return None


def clean_player_name(raw: str) -> str:
    """Clean player name from various formats."""
    if not raw:
        return ""

    s = str(raw).strip()

    # Remove leading rank number like "7. " or "18. "
    s = re.sub(r'^\d+\.\s*', '', s)

    # Handle "LastName, FirstName" format
    if ',' in s:
        parts = s.split(',', 1)
        if len(parts) == 2:
            last = parts[0].strip()
            rest = parts[1].strip()
            # rest might include team like "Russell SEA" — just take first word(s)
            first = rest.split()[0] if rest else ""
            return f"{first} {last}".strip()

    # Handle "FirstName LastName, TEAM" format
    s = re.sub(r',\s*[A-Z]{2,3}$', '', s)

    return s.strip()


def parse_position(raw) -> str:
    """Extract position from various formats like 'RB5', ' RB', 'QB', etc."""
    if not raw:
        return None
    s = str(raw).strip().upper()
    # Match position at start
    match = re.match(r'(QB|RB|WR|TE|DEF|DST|K|D/ST)', s)
    if match:
        pos = match.group(1)
        if pos in ("DST", "D/ST"):
            pos = "DEF"
        return pos
    return None


# ============================================================
# FILE PARSERS (one per format)
# ============================================================

def parse_master_draft_list(filepath: str, year: int) -> list:
    """Parse 2011-2019 format: Master Draft List tab."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Find the right sheet
    sheet_name = None
    for name in wb.sheetnames:
        if "master draft" in name.lower():
            sheet_name = name
            break

    if not sheet_name:
        print(f"  ❌ No Master Draft sheet found in {filepath}")
        wb.close()
        return []

    ws = wb[sheet_name]
    picks = []
    current_round = 0
    pick_in_round = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        overall = row[0]
        owner_raw = str(row[1]).strip() if row[1] else ""

        # Skip round headers
        if not overall or not isinstance(overall, (int, float)):
            if "round" in owner_raw.lower():
                current_round += 1
                pick_in_round = 0
            continue

        overall = int(overall)
        current_round = ((overall - 1) // 12) + 1
        pick_in_round = ((overall - 1) % 12) + 1

        # Player name — might be in col 2, or split across cols 2+3
        player_raw = ""
        if year == 2018:
            # 2018 format: First Name, Last Name in cols 2, 3
            first = str(row[2]).strip() if row[2] else ""
            last = str(row[3]).strip() if row[3] else ""
            player_raw = f"{first} {last}".strip()
            keeper_raw = row[4] if len(row) > 4 else None
            position = None  # 2018 doesn't have position in master list
        elif year == 2019:
            # 2019: player in col 2, position col 3, team col 4
            player_raw = str(row[2]).strip() if row[2] else ""
            position = parse_position(row[3])
            keeper_raw = None  # Check if keeper info exists
        else:
            # 2011-2017: player in col 2, position in col 3
            player_raw = str(row[2]).strip() if row[2] else ""
            pos_raw = row[3] if len(row) > 3 else None
            position = parse_position(pos_raw)
            # Check for keeper marker in later columns
            keeper_raw = None
            for col_val in row[4:8] if len(row) > 4 else []:
                if col_val and 'K' in str(col_val).upper():
                    k = parse_keeper(col_val)
                    if k is not None:
                        keeper_raw = str(col_val)
                        break

        player_name = clean_player_name(player_raw)
        traded_from = parse_traded_from(owner_raw)
        owner = normalize_owner(owner_raw, year)
        original_owner = normalize_owner(traded_from, year) if traded_from else owner
        keeper_year = parse_keeper(keeper_raw) if 'keeper_raw' in dir() and keeper_raw else None

        if not player_name:
            continue

        picks.append({
            "overall": overall,
            "round": current_round,
            "pick_in_round": pick_in_round,
            "owner": owner,
            "original_owner": original_owner,
            "player_name": player_name,
            "position": position if 'position' in dir() else None,
            "keeper_year": keeper_year,
        })

    wb.close()
    return picks


def parse_fandraft_by_round(filepath: str, year: int, sheet_name: str) -> list:
    """Parse 2020+ FanDraft format: Draft by Round tab."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        print(f"  ❌ Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        wb.close()
        return []

    ws = wb[sheet_name]
    picks = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        try:
            overall = int(float(str(row[0])))
        except (ValueError, TypeError):
            continue

        if year == 2020:
            # Format: Pick, Team, Player, PlayerTeam, Position, Keeper
            team_raw = str(row[1]).strip() if row[1] else ""
            player_raw = str(row[2]).strip() if row[2] else ""
            position = parse_position(row[4])
            keeper_raw = row[5] if len(row) > 5 else None
            owner = normalize_owner(team_raw, year)
            original_owner = owner
        elif year == 2021:
            # Format: Overall, Round/Pick, Team, Player, PlayerTeam, Position, Keeper
            owner_raw = str(row[2]).strip() if row[2] else ""
            player_raw = str(row[3]).strip() if row[3] else ""
            position = parse_position(row[5])
            keeper_raw = row[6] if len(row) > 6 else None
            traded_from = parse_traded_from(owner_raw)
            owner = normalize_owner(owner_raw, year)
            original_owner = normalize_owner(traded_from, year) if traded_from else owner
        elif year == 2022:
            # Format: Pick, Round/Pick, Team, Owner, Player, PlayerTeam, Position, KeeperStatus
            owner_raw = str(row[3]).strip() if row[3] else ""
            player_raw = str(row[4]).strip() if row[4] else ""
            position = parse_position(row[6])
            keeper_raw = row[7] if len(row) > 7 else None
            owner = normalize_owner(owner_raw, year)
            original_owner = owner
        else:
            continue

        round_num = ((overall - 1) // 12) + 1
        pick_in_round = ((overall - 1) % 12) + 1

        player_name = clean_player_name(player_raw)
        keeper_year = parse_keeper(keeper_raw)

        if not player_name:
            continue

        picks.append({
            "overall": overall,
            "round": round_num,
            "pick_in_round": pick_in_round,
            "owner": owner,
            "original_owner": original_owner,
            "player_name": player_name,
            "position": position,
            "keeper_year": keeper_year,
        })

    wb.close()
    return picks


def parse_fandraft_csv(filepath: str, year: int) -> list:
    """Parse FanDraft CSV export: Round/Pick, Team, Amount, Player, PlayerTeam, PlayerPosition."""
    import csv

    picks = []
    with open(filepath, 'r') as f:
        reader = csv.DictReader(f)
        overall = 0

        for row in reader:
            overall += 1
            round_pick = row.get("Round/Pick", "")
            team_raw = row.get("Team", "")
            player_raw = row.get("Player", "")
            position = parse_position(row.get("Player Position", ""))

            # Parse round/pick from "1.1" format
            parts = round_pick.split(".")
            if len(parts) == 2:
                try:
                    round_num = int(parts[0])
                    pick_in_round = int(parts[1])
                except ValueError:
                    continue
            else:
                round_num = ((overall - 1) // 12) + 1
                pick_in_round = ((overall - 1) % 12) + 1

            owner = normalize_owner(team_raw, year)
            player_name = clean_player_name(player_raw)

            if not player_name:
                continue

            picks.append({
                "overall": overall,
                "round": round_num,
                "pick_in_round": pick_in_round,
                "owner": owner,
                "original_owner": owner,
                "player_name": player_name,
                "position": position,
                "keeper_year": None,  # Keeper info not in FanDraft CSV
            })

    return picks


# ============================================================
# SUPABASE HELPERS
# ============================================================

_owner_cache: dict = {}

def get_owner_id(name: str) -> str:
    """Get owner UUID from Supabase by name."""
    if name in _owner_cache:
        return _owner_cache[name]

    resp = supabase.table("owners").select("id").eq("name", name).execute()
    if resp.data:
        _owner_cache[name] = resp.data[0]["id"]
        return _owner_cache[name]
    return None


_player_cache: dict = {}

def get_or_create_player(name: str, position: str) -> int:
    """Get or create a player, return their ID."""
    cache_key = name.lower()
    if cache_key in _player_cache:
        return _player_cache[cache_key]

    # Try to find existing
    resp = supabase.table("players").select("id").ilike("name", name).execute()
    if resp.data:
        _player_cache[cache_key] = resp.data[0]["id"]
        return _player_cache[cache_key]

    # Create new player
    pos = position or "RB"  # Default if unknown

    # Try to detect DEF from name
    if any(team in name.upper() for team in ["D/ST", "DST", "DEFENSE"]) or \
       re.match(r'^[A-Z][a-z]+ [A-Z][a-z]+$', name) is None and \
       any(t in name for t in ["49ers", "Cowboys", "Bills", "Jets", "Browns", "Ravens",
                                 "Jaguars", "Rams", "Eagles", "Chargers", "Steelers",
                                 "Vikings", "Patriots", "Panthers", "Texans", "Broncos",
                                 "Saints", "Bears", "Packers", "Seahawks", "Cardinals",
                                 "Lions", "Falcons", "Giants", "Raiders", "Colts",
                                 "Titans", "Chiefs", "Bengals", "Dolphins", "Commanders",
                                 "Buccaneers"]):
        pos = "DEF"

    # Also check for team abbreviation patterns in name
    team_abbrs = ["Jacksonville", "Los Angeles", "New England", "New Orleans",
                  "Philadelphia", "Pittsburgh", "Minnesota", "Houston", "Denver",
                  "Baltimore", "Carolina", "Seattle"]
    for abbr in team_abbrs:
        if abbr.lower() in name.lower():
            pos = "DEF"
            break

    resp = supabase.table("players").insert({
        "name": name,
        "position": pos,
        "is_active": False,  # Historical player
    }).execute()

    if resp.data:
        _player_cache[cache_key] = resp.data[0]["id"]
        return _player_cache[cache_key]

    raise Exception(f"Failed to create player: {name}")


def get_or_create_season(year: int) -> int:
    """Get or create a season, return its ID."""
    resp = supabase.table("seasons").select("id").eq("year", year).execute()
    if resp.data:
        return resp.data[0]["id"]

    resp = supabase.table("seasons").insert({
        "year": year,
        "draft_status": "complete",
        "draft_order": [],
    }).execute()
    return resp.data[0]["id"]


def import_picks(year: int, picks: list):
    """Import parsed picks into Supabase."""
    season_id = get_or_create_season(year)

    # Check if picks already exist for this season
    existing = supabase.table("draft_picks").select("id").eq("season_id", season_id).limit(1).execute()
    if existing.data:
        print(f"  ⚠️  Draft picks already exist for {year} — skipping (delete first to re-import)")
        return

    success = 0
    errors = 0

    for pick in picks:
        owner_id = get_owner_id(pick["owner"])
        original_owner_id = get_owner_id(pick["original_owner"])

        if not owner_id:
            print(f"  ❌ No owner ID for '{pick['owner']}' (pick #{pick['overall']})")
            errors += 1
            continue

        if not original_owner_id:
            original_owner_id = owner_id

        player_id = get_or_create_player(pick["player_name"], pick.get("position"))

        resp = supabase.table("draft_picks").insert({
            "season_id": season_id,
            "round": pick["round"],
            "pick_in_round": pick["pick_in_round"],
            "overall_pick": pick["overall"],
            "original_owner_id": original_owner_id,
            "current_owner_id": owner_id,
            "player_id": player_id,
            "is_keeper": pick["keeper_year"] is not None,
            "keeper_year": pick["keeper_year"],
            "picked_at": f"{year}-09-01T00:00:00Z",  # Approximate draft date
        }).execute()

        if resp.data:
            success += 1
        else:
            print(f"  ❌ Failed to insert pick #{pick['overall']}: {pick['player_name']}")
            errors += 1

    print(f"  ✅ {success} picks imported, {errors} errors")


# ============================================================
# MAIN
# ============================================================

DRAFT_FILES = [
    # (filename, year, parser, sheet_name)
    ("Bust A Nut League Draft Sheet.xlsx", 2011, "master", None),
    ("2013 Bust A Nut League Draft Sheet.xlsx", 2013, "master", None),
    ("2014 Bust A Nut League Draft Sheet.xlsx", 2014, "master", None),
    ("2015 BANL Draft Sheet.xlsx", 2015, "master", None),
    ("2016 BANL Draft Sheet.xlsx", 2016, "master", None),
    ("2017 BANL Draft Sheet.xlsx", 2017, "master", None),
    ("2018 BANL Draft Sheet.xlsx", 2018, "master", None),
    ("2019 BANL Draft Sheet.xlsx", 2019, "master", None),
    ("BANL 2020 Draft Results.xlsx", 2020, "fandraft", "Draft by Round"),
    ("2021-2022 BANL Draft.xlsx", 2021, "fandraft", "Draft by Round"),
    ("2022 Draft Results.xlsx", 2022, "fandraft", "2022 Draft by Round wKeepers"),
    ("draft-summary-BANL 2022-2026-03-19.csv", 2022, "csv", None),
    ("draft-summary-BANL 2023-2026-03-19.csv", 2023, "csv", None),
    ("draft-summary-BANL 2024-2026-03-19.csv", 2024, "csv", None),
    ("draft-summary-BANL 2025-2026-03-19.csv", 2025, "csv", None),
]


def main():
    print("🏈 Draft History Importer")
    print("=" * 50)

    for filename, year, parser_type, sheet_name in DRAFT_FILES:
        filepath = os.path.join(BASE_DIR, filename)

        if not os.path.exists(filepath):
            print(f"\n⚠️  {filename} not found — skipping")
            continue

        print(f"\n📄 {year}: {filename}")

        if parser_type == "master":
            picks = parse_master_draft_list(filepath, year)
        elif parser_type == "fandraft":
            picks = parse_fandraft_by_round(filepath, year, sheet_name)
        elif parser_type == "csv":
            picks = parse_fandraft_csv(filepath, year)
        else:
            print(f"  ❌ Unknown parser: {parser_type}")
            continue

        if not picks:
            print(f"  ❌ No picks parsed")
            continue

        print(f"  📊 Parsed {len(picks)} picks")

        # Show first 3 for verification
        for p in picks[:3]:
            print(f"     #{p['overall']:3d} R{p['round']}.{p['pick_in_round']:02d} | {p['owner']:20s} | {p['player_name']:25s} | {str(p.get('position') or '?'):3s} | {'K'+str(p['keeper_year']) if p['keeper_year'] else ''}")

        import_picks(year, picks)

    print("\n✅ Done!")


if __name__ == "__main__":
    main()
